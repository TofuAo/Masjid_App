"""
Excel Payment Calculator Generator
Creates an Excel file to calculate upfront payment plans (per month/per year)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_payment_calculator():
    """Create an Excel file for payment calculation with upfront plans"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Calculator"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="1F4E78")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title
    ws['A1'] = "VPS PREMIUM + WHM + CPANEL - PAYMENT CALCULATOR"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = center_align
    
    # Subtitle
    ws['A2'] = "Upfront Payment (Lump Sum) - Monthly & Yearly Cost Calculations"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = center_align
    
    # Empty row
    ws.row_dimensions[3].height = 5
    
    # Pricing Configuration Section
    row = 4
    ws.cell(row=row, column=1).value = "PRICING CONFIGURATION"
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{row}:G{row}')
    
    # Add-ons and Tax Settings
    row += 1
    ws.cell(row=row, column=1).value = "Add-on Services (Monthly):"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "Data Backup"
    ws.cell(row=row, column=3).value = 58.90
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=3).alignment = right_align
    
    row += 1
    ws.cell(row=row, column=1).value = "Service Tax:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = "8%"
    ws.cell(row=row, column=3).value = 0.08
    ws.cell(row=row, column=3).number_format = '0.00%'
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=3).alignment = right_align
    
    # Store values for calculations
    backup_rate = 58.90
    tax_rate = 0.08
    
    # Empty row
    row += 2
    
    # Plan Configuration Section
    ws.cell(row=row, column=1).value = "PLAN CONFIGURATION"
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{row}:G{row}')
    
    # Headers
    row += 1
    headers = ['Plan Name', 'Duration (Months)', 'Base Rate/mo (RM)', 'Sub-total/mo (RM)', 'Sub-total (RM)', 'Tax 8% (RM)', 'Total Upfront (RM)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # VPS Premium + WHM + cPanel Plans (Monthly Rate × Duration = Upfront Payment)
    # Format: [Plan Name, Duration (Months), Monthly Rate (RM)]
    plans = [
        ['1 Month Plan', 1, 194.90],
        ['3 Month Plan', 3, 189.90],
        ['6 Month Plan', 6, 184.90],
        ['12 Month Plan', 12, 179.90],
        ['24 Month Plan', 24, 169.90],
        ['36 Month Plan', 36, 162.90],
    ]
    
    # Add plan data with calculated static values
    plan_start_row = row + 1
    for plan_idx, plan in enumerate(plans):
        plan_row = plan_start_row + plan_idx
        
        # Calculate values
        base_rate = plan[2]
        duration = plan[1]
        subtotal_month = base_rate + backup_rate
        subtotal_duration = subtotal_month * duration
        tax = subtotal_duration * tax_rate
        total_upfront = subtotal_duration + tax
        
        # Plan Name
        ws.cell(row=plan_row, column=1).value = plan[0]
        ws.cell(row=plan_row, column=1).border = border
        ws.cell(row=plan_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        # Duration
        ws.cell(row=plan_row, column=2).value = duration
        ws.cell(row=plan_row, column=2).border = border
        ws.cell(row=plan_row, column=2).alignment = center_align
        
        # Base Monthly Rate (VPS only)
        ws.cell(row=plan_row, column=3).value = base_rate
        ws.cell(row=plan_row, column=3).border = border
        ws.cell(row=plan_row, column=3).alignment = right_align
        ws.cell(row=plan_row, column=3).number_format = '#,##0.00'
        
        # Sub-total per Month (Base + Backup) - STATIC VALUE
        subtotal_month_cell = ws.cell(row=plan_row, column=4)
        subtotal_month_cell.value = subtotal_month
        subtotal_month_cell.border = border
        subtotal_month_cell.alignment = right_align
        subtotal_month_cell.number_format = '#,##0.00'
        
        # Sub-total for Duration - STATIC VALUE
        subtotal_duration_cell = ws.cell(row=plan_row, column=5)
        subtotal_duration_cell.value = subtotal_duration
        subtotal_duration_cell.border = border
        subtotal_duration_cell.alignment = right_align
        subtotal_duration_cell.number_format = '#,##0.00'
        
        # Service Tax - STATIC VALUE
        tax_cell = ws.cell(row=plan_row, column=6)
        tax_cell.value = tax
        tax_cell.border = border
        tax_cell.alignment = right_align
        tax_cell.number_format = '#,##0.00'
        
        # Total Upfront Payment - STATIC VALUE
        upfront_cell = ws.cell(row=plan_row, column=7)
        upfront_cell.value = total_upfront
        upfront_cell.border = border
        upfront_cell.alignment = right_align
        upfront_cell.number_format = '#,##0.00'
        upfront_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        upfront_cell.font = Font(bold=True)
    
    # Empty row
    row = plan_row + 2
    
    # Comparison Section
    ws.cell(row=row, column=1).value = "PLAN COMPARISON SUMMARY"
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    comparison_headers = ['Plan Name', 'Duration', 'Sub-total/mo (RM)', 'Sub-total (RM)', 'Tax 8% (RM)', 'Total Upfront (RM)', 'Monthly Avg (RM)']
    for col, header in enumerate(comparison_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Comparison data (static values from plan configuration)
    for plan_idx, plan in enumerate(plans):
        comp_row = row + 1 + plan_idx
        plan_data_row = plan_start_row + plan_idx
        
        # Calculate values
        base_rate = plan[2]
        duration = plan[1]
        subtotal_month = base_rate + backup_rate
        subtotal_duration = subtotal_month * duration
        tax = subtotal_duration * tax_rate
        total_upfront = subtotal_duration + tax
        monthly_avg = total_upfront / duration
        
        # Plan Name
        ws.cell(row=comp_row, column=1).value = plan[0]
        ws.cell(row=comp_row, column=1).border = border
        
        # Duration
        ws.cell(row=comp_row, column=2).value = duration
        ws.cell(row=comp_row, column=2).border = border
        ws.cell(row=comp_row, column=2).alignment = center_align
        
        # Sub-total per Month - STATIC VALUE
        ws.cell(row=comp_row, column=3).value = subtotal_month
        ws.cell(row=comp_row, column=3).border = border
        ws.cell(row=comp_row, column=3).alignment = right_align
        ws.cell(row=comp_row, column=3).number_format = '#,##0.00'
        
        # Sub-total - STATIC VALUE
        ws.cell(row=comp_row, column=4).value = subtotal_duration
        ws.cell(row=comp_row, column=4).border = border
        ws.cell(row=comp_row, column=4).alignment = right_align
        ws.cell(row=comp_row, column=4).number_format = '#,##0.00'
        
        # Tax - STATIC VALUE
        ws.cell(row=comp_row, column=5).value = tax
        ws.cell(row=comp_row, column=5).border = border
        ws.cell(row=comp_row, column=5).alignment = right_align
        ws.cell(row=comp_row, column=5).number_format = '#,##0.00'
        
        # Total Upfront Payment - STATIC VALUE
        ws.cell(row=comp_row, column=6).value = total_upfront
        ws.cell(row=comp_row, column=6).border = border
        ws.cell(row=comp_row, column=6).alignment = right_align
        ws.cell(row=comp_row, column=6).number_format = '#,##0.00'
        ws.cell(row=comp_row, column=6).fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        ws.cell(row=comp_row, column=6).font = Font(bold=True)
        
        # Monthly Average - STATIC VALUE
        ws.cell(row=comp_row, column=7).value = monthly_avg
        ws.cell(row=comp_row, column=7).border = border
        ws.cell(row=comp_row, column=7).alignment = right_align
        ws.cell(row=comp_row, column=7).number_format = '#,##0.00'
    
    # Empty row
    row = comp_row + 2
    
    # Savings Analysis Section
    ws.cell(row=row, column=1).value = "SAVINGS ANALYSIS"
    ws.cell(row=row, column=1).font = subheader_font
    ws.cell(row=row, column=1).fill = subheader_fill
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    savings_headers = ['Plan', 'vs 1 Month Plan', 'Savings (RM)', 'Savings %', 'Best Value']
    for col, header in enumerate(savings_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Calculate savings compared to 1 month plan (based on Total Upfront Payment) - STATIC VALUES
    savings_start_row = row + 1
    
    # Calculate all plan totals first
    plan_totals = []
    for plan in plans:
        base_rate = plan[2]
        duration = plan[1]
        subtotal_month = base_rate + backup_rate
        subtotal_duration = subtotal_month * duration
        tax = subtotal_duration * tax_rate
        total_upfront = subtotal_duration + tax
        plan_totals.append(total_upfront)
    
    # Find best value (highest savings percentage)
    savings_percentages = []
    for idx in range(len(plans)):
        if idx == 0:
            savings_percentages.append(0)
        else:
            one_month_total = plan_totals[0]
            plan_total = plan_totals[idx]
            duration = plans[idx][1]
            savings = (one_month_total * duration) - plan_total
            savings_pct = (savings / (one_month_total * duration)) * 100 if (one_month_total * duration) > 0 else 0
            savings_percentages.append(savings_pct)
    
    max_savings_pct = max(savings_percentages) if savings_percentages else 0
    
    for idx, plan in enumerate(plans):
        savings_row = savings_start_row + idx
        
        # Plan Name
        ws.cell(row=savings_row, column=1).value = plan[0]
        ws.cell(row=savings_row, column=1).border = border
        
        # Comparison text
        if idx == 0:
            ws.cell(row=savings_row, column=2).value = "Base Plan"
        else:
            ws.cell(row=savings_row, column=2).value = "vs 1 Month"
        ws.cell(row=savings_row, column=2).border = border
        ws.cell(row=savings_row, column=2).alignment = center_align
        
        # Savings calculation - STATIC VALUE
        if idx == 0:
            savings_amount = 0
        else:
            one_month_total = plan_totals[0]
            plan_total = plan_totals[idx]
            duration = plan[1]
            savings_amount = (one_month_total * duration) - plan_total
        
        ws.cell(row=savings_row, column=3).value = savings_amount
        ws.cell(row=savings_row, column=3).border = border
        ws.cell(row=savings_row, column=3).alignment = right_align
        ws.cell(row=savings_row, column=3).number_format = '#,##0.00'
        
        # Savings percentage - STATIC VALUE
        savings_pct = savings_percentages[idx]
        if idx == 0:
            ws.cell(row=savings_row, column=4).value = "0%"
        else:
            ws.cell(row=savings_row, column=4).value = savings_pct / 100
            ws.cell(row=savings_row, column=4).number_format = '0.00"%"'
        ws.cell(row=savings_row, column=4).border = border
        ws.cell(row=savings_row, column=4).alignment = center_align
        
        # Best value indicator - STATIC VALUE
        if idx > 0 and savings_pct == max_savings_pct:
            ws.cell(row=savings_row, column=5).value = "BEST"
        else:
            ws.cell(row=savings_row, column=5).value = ""
        ws.cell(row=savings_row, column=5).border = border
        ws.cell(row=savings_row, column=5).alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    
    # Add footer note
    footer_row = savings_start_row + len(plans) + 2
    ws.cell(row=footer_row, column=1).value = "Note: All payments are paid upfront (lump sum). Total Upfront = (Base + Add-ons) × Duration × (1 + Tax). All values are static data."
    ws.cell(row=footer_row, column=1).font = Font(size=9, italic=True, color="666666")
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    
    # Add timestamp
    footer_row += 1
    ws.cell(row=footer_row, column=1).value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.cell(row=footer_row, column=1).font = Font(size=8, italic=True, color="999999")
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    
    # Save the file
    filename = "Payment_Calculator_Upfront.xlsx"
    wb.save(filename)
    print(f"Excel file created successfully: {filename}")
    print(f"\nFile includes:")
    print("  - Plan Configuration (static data)")
    print("  - Plan Comparison (static data)")
    print("  - Savings Analysis (static data)")
    
    print(f"\nVPS Premium Pricing Plans (with Data Backup + 8% Tax):")
    base_rate = 194.90
    backup = 58.90
    tax_rate = 0.08
    for plan in plans:
        subtotal_month = plan[2] + backup
        subtotal_duration = subtotal_month * plan[1]
        tax = subtotal_duration * tax_rate
        total_upfront = subtotal_duration + tax
        print(f"  - {plan[0]}: Base RM{plan[2]:.2f}/mo + Backup RM{backup:.2f}/mo = RM{subtotal_month:.2f}/mo")
        print(f"    {plan[1]} months: RM{subtotal_duration:,.2f} + Tax RM{tax:,.2f} = RM{total_upfront:,.2f} upfront")
    print(f"\nAll values are static data (not formulas).")

if __name__ == "__main__":
    create_payment_calculator()
