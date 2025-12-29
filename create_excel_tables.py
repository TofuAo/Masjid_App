import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "VPS Comparison"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="FFFFFF")
section_font = Font(bold=True, size=12)
highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

current_row = 1

def add_section_title(title, row):
    """Add a section title"""
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = title_font
    cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return row + 1

def add_table(data, start_row, has_header=True):
    """Add a table with data"""
    row = start_row
    for i, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = border
            
            if has_header and i == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
            # Highlight TOTAL DUE TODAY rows
            if isinstance(value, str) and 'TOTAL DUE TODAY' in value.upper():
                for c in range(1, len(row_data) + 1):
                    ws.cell(row=row, column=c).fill = highlight_fill
                    ws.cell(row=row, column=c).font = Font(bold=True)
        
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(data[0]) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return row + 2  # Return next row with spacing

# Quick Reference
current_row = add_section_title("QUICK REFERENCE", current_row)
quick_ref = [
    ["Product", "Value VPS", "Pro VPS"],
    ["Product Name", "Automated KVM Windows VPS Value with Plesk", "Automated KVM Windows VPS Pro with Plesk"],
    ["Best Price 36 months", "RM 98.90 per month", "RM 127.90 per month"],
    ["Best Total 36 months", "RM 3560.40", "RM 4604.40"],
    ["Monthly Difference", "Base Price", "RM 39.00 more for Pro"]
]
current_row = add_table(quick_ref, current_row)

# Section 1: Product Specifications
current_row = add_section_title("SECTION 1: PRODUCT SPECIFICATIONS", current_row)
specs = [
    ["Specification", "Value VPS", "Pro VPS"],
    ["vCPUs", "2", "4"],
    ["RAM", "2 GB", "4 GB"],
    ["SSD Storage", "50 GB", "100 GB"],
    ["IP Address", "1", "1"],
    ["Network Speed", "100 Mbps", "100 Mbps"],
    ["Data Transfer", "Unlimited", "Unlimited"],
    ["Hypervisor", "KVM", "KVM"],
    ["Operating System", "Windows 2022 Plesk 18 Obsidian", "Windows 2022 Plesk 18 Obsidian"],
    ["Control Panel", "Plesk 10 Domains Web Admin Edition", "Plesk 10 Domains Web Admin Edition"],
    ["Managed Services", "BASIC Silver Managed Service Included", "BASIC Silver Managed Service Included"],
    ["Setup Fee", "RM 0.00", "RM 0.00"]
]
current_row = add_table(specs, current_row)

# Section 2: Pricing Monthly Rates
current_row = add_section_title("SECTION 2: PRICING MONTHLY RATES", current_row)
pricing = [
    ["Billing Cycle", "Value VPS per month", "Pro VPS per month", "Savings"],
    ["1 Month", "RM 125.90", "RM 164.90", "0 percent"],
    ["3 Months", "RM 119.90", "RM 154.90", "10 percent off"],
    ["6 Months", "RM 117.90", "RM 150.90", "14 percent off"],
    ["12 Months 1 Year", "RM 113.90", "RM 144.90", "21 percent off"],
    ["24 Months 2 Years", "RM 106.90", "RM 137.90", "34 percent off Value or 29 percent off Pro"],
    ["36 Months 3 Years", "RM 98.90", "RM 127.90", "49 percent off Value or 40 percent off Pro"]
]
current_row = add_table(pricing, current_row)

# Section 3: Total Amount to Pay
current_row = add_section_title("SECTION 3: TOTAL AMOUNT TO PAY ALL PLANS", current_row)
total_amounts = [
    ["Billing Cycle", "Value VPS Total Due", "Pro VPS Total Due", "Difference"],
    ["1 Month Plan", "RM 125.90", "RM 164.90", "RM 39.00"],
    ["3 Months Plan", "RM 359.70", "RM 464.70", "RM 105.00"],
    ["6 Months Plan", "RM 707.40", "RM 905.40", "RM 198.00"],
    ["12 Months Plan 1 Year", "RM 1366.80", "RM 1738.80", "RM 372.00"],
    ["24 Months Plan 2 Years", "RM 2565.60", "RM 3309.60", "RM 744.00"],
    ["36 Months Plan 3 Years", "RM 3560.40", "RM 4604.40", "RM 1044.00"]
]
current_row = add_table(total_amounts, current_row)

# Add note
note_cell = ws.cell(row=current_row, column=1)
note_cell.value = "NOTE: All amounts shown are TOTAL DUE TODAY (full payment upfront). Setup Fee: RM 0.00 for both plans."
note_cell.font = Font(italic=True)
ws.merge_cells(f'A{current_row}:D{current_row}')
current_row += 3

# Section 4: Payment Breakdown by Plan
current_row = add_section_title("SECTION 4: PAYMENT BREAKDOWN BY PLAN", current_row)

# Plan 1 Month
plan1 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 125.90", "RM 164.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["TOTAL DUE TODAY", "RM 125.90", "RM 164.90"]
]
current_row = add_table(plan1, current_row)

# Plan 3 Months
plan3 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 119.90", "RM 154.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["Total Amount 3 months", "RM 359.70", "RM 464.70"],
    ["TOTAL DUE TODAY", "RM 359.70", "RM 464.70"]
]
current_row = add_table(plan3, current_row)

# Plan 6 Months
plan6 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 117.90", "RM 150.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["Total Amount 6 months", "RM 707.40", "RM 905.40"],
    ["TOTAL DUE TODAY", "RM 707.40", "RM 905.40"]
]
current_row = add_table(plan6, current_row)

# Plan 12 Months
plan12 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 113.90", "RM 144.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["Total Amount 12 months", "RM 1366.80", "RM 1738.80"],
    ["TOTAL DUE TODAY", "RM 1366.80", "RM 1738.80"]
]
current_row = add_table(plan12, current_row)

# Plan 24 Months
plan24 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 106.90", "RM 137.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["Total Amount 24 months", "RM 2565.60", "RM 3309.60"],
    ["TOTAL DUE TODAY", "RM 2565.60", "RM 3309.60"]
]
current_row = add_table(plan24, current_row)

# Plan 36 Months
plan36 = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Rate", "RM 98.90", "RM 127.90"],
    ["Setup Fee", "RM 0.00", "RM 0.00"],
    ["Total Amount 36 months", "RM 3560.40", "RM 4604.40"],
    ["TOTAL DUE TODAY", "RM 3560.40", "RM 4604.40"],
    ["Savings", "49 percent off Value", "40 percent off Pro"]
]
current_row = add_table(plan36, current_row)

# Section 5: Quick Price Comparison
current_row = add_section_title("SECTION 5: QUICK PRICE COMPARISON", current_row)
quick_price = [
    ["Scenario", "Value VPS", "Pro VPS"],
    ["Cheapest Option 36 months", "RM 3560.40", "RM 4604.40"],
    ["Most Popular 12 months", "RM 1366.80", "RM 1738.80"],
    ["Short Term 3 months", "RM 359.70", "RM 464.70"],
    ["Monthly 1 month", "RM 125.90", "RM 164.90"]
]
current_row = add_table(quick_price, current_row)

# Section 6: Cost Analysis
current_row = add_section_title("SECTION 6: COST ANALYSIS", current_row)
cost_analysis = [
    ["Item", "Value VPS", "Pro VPS"],
    ["Monthly Price Difference 1 month", "Base RM 125.90", "Base RM 164.90"],
    ["Extra Cost for Pro per month", "", "RM 39.00 more"],
    ["Extra Cost for Pro per year", "", "RM 372.00 more"],
    ["Extra Cost for Pro 3 years", "", "RM 1044.00 more"],
    ["Pro is More Expensive By", "Base Price", "31.0 percent more expensive"],
    ["3 Year Total Cost", "RM 3560.40", "RM 4604.40"]
]
current_row = add_table(cost_analysis, current_row)

# Section 7: Pros and Cons
current_row = add_section_title("SECTION 7: PROS AND CONS", current_row)

# VPS Value Pros
value_pros = [
    ["VPS VALUE PROS"],
    ["Lower initial cost RM 125.90 per month"],
    ["Best long term savings 49 percent off at 36 months"],
    ["Good for small to medium websites"],
    ["Perfect for development and testing"],
    ["Lower total cost of ownership"],
    ["Adequate for basic applications"]
]
current_row = add_table(value_pros, current_row, has_header=False)

# VPS Value Cons
value_cons = [
    ["VPS VALUE CONS"],
    ["Only 2 vCPUs limited processing"],
    ["2GB RAM may be insufficient for heavy apps"],
    ["50GB storage may fill up quickly"],
    ["Not suitable for resource intensive apps"],
    ["May have performance issues under heavy load"],
    ["Limited scalability"]
]
current_row = add_table(value_cons, current_row, has_header=False)

# VPS Pro Pros
pro_pros = [
    ["VPS PRO PROS"],
    ["Double processing power 4 vCPUs"],
    ["Double RAM 4GB for better performance"],
    ["Double storage 100GB for more data"],
    ["Better for production environments"],
    ["Can handle more concurrent users"],
    ["Suitable for resource intensive apps"],
    ["Better performance under heavy load"],
    ["More room for growth"]
]
current_row = add_table(pro_pros, current_row, has_header=False)

# VPS Pro Cons
pro_cons = [
    ["VPS PRO CONS"],
    ["Higher initial cost RM 164.90 per month"],
    ["Lower savings 40 percent vs 49 percent at 36 months"],
    ["May be overkill for simple websites"],
    ["Higher total cost RM 1044 more over 3 years"],
    ["More expensive upfront"]
]
current_row = add_table(pro_cons, current_row, has_header=False)

# Section 8: Recommendations
current_row = add_section_title("SECTION 8: RECOMMENDATIONS BY USE CASE", current_row)
recommendations = [
    ["Use Case", "Recommended Plan", "Reason"],
    ["Small Website Blog", "Value VPS", "Sufficient resources at lower cost"],
    ["Development Testing", "Value VPS", "Cost effective for testing"],
    ["Small Business Website", "Value VPS", "Good balance of cost and performance"],
    ["E commerce Site", "Pro VPS", "Better performance needed"],
    ["Resource Intensive Apps", "Pro VPS", "Requires more power"],
    ["Multiple Websites", "Pro VPS", "Better capacity needed"]
]
current_row = add_table(recommendations, current_row)

# Section 9: Feature Comparison
current_row = add_section_title("SECTION 9: FEATURE COMPARISON", current_row)
features = [
    ["Feature", "Value VPS", "Pro VPS"],
    ["Processing Power vCPUs", "2 Lower", "4 Higher 2x"],
    ["Memory RAM", "2 GB Lower", "4 GB Higher 2x"],
    ["Storage SSD", "50 GB Lower", "100 GB Higher 2x"],
    ["Network Speed", "100 Mbps Same", "100 Mbps Same"],
    ["Data Transfer", "Unlimited Same", "Unlimited Same"],
    ["Managed Services", "Included Same", "Included Same"],
    ["Control Panel", "Plesk Same", "Plesk Same"],
    ["Setup Fee", "Free Same", "Free Same"]
]
current_row = add_table(features, current_row)

# Important Notes
current_row = add_section_title("IMPORTANT NOTES", current_row)
notes = [
    ["All prices are in Malaysian Ringgit RM"],
    ["Setup fee is RM 0.00 for both plans"],
    ["Payment is required upfront full amount"],
    ["Prices shown are per month rates for the billing cycle"],
    ["Longer commitments equal Better savings"],
    ["Best value 36 month plan for both products"]
]
current_row = add_table(notes, current_row, has_header=False)

# Save the workbook
wb.save("VPS_Product_Comparison.xlsx")
print("Excel file created successfully: VPS_Product_Comparison.xlsx")

